import gspread
import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, colors

with open(os.path.join(os.path.dirname(__file__),"classesdone.json"), 'r') as f:
    data = json.load(f)
cred_name = os.path.join(os.path.dirname(__file__),'drive_credentials.json')

sa = gspread.service_account(filename=cred_name)
sh = sa.open("Nida's Classes Database Auto")

wks = sh.worksheet("Classes Data")

row=1

for n,day in enumerate(data):
    wks.update('A'+str(row),data[n]["day"])
    wks.format('A'+str(row), {"horizontalAlignment": "CENTER","textFormat": {"fontSize": 12,"bold": True}})
    wks.update('A'+str(row),data[n]["date"])
    wks.format('A'+str(row), {"horizontalAlignment": "CENTER","textFormat": {"fontSize": 12,"bold": True}})

#    if data[n]["day"]=="Sunday":
#        wks['A'+str(row)].font=Font(bold=True, size=12, color='FF0000')
#        wks['B'+str(row)].font=Font(bold=True, size=12, color='FF0000')

    m=0
    if len(data[n]["classes"]) > 0:
        for m,session in enumerate(data[n]["classes"]):
            wks.update('C'+str(row+m),session["section"])
            wks.update('D'+str(row+m),session["cond"])
            wks.update('E'+str(row+m),session["sch_time"])
            wks.update('F'+str(row+m),session["act_time"])
            wks.update('G'+str(row+m),session["hours"])
            if session["hours"] != "  Hours  ":
                if session["hours"]!= "":
                    hrs_red=session["hours"].split(":")
                    hrs_red=round((int(hrs_red[0])+(int(hrs_red[1])/60)),3)
                    wks.update('H'+str(row+m),hrs_red)
            else:
                wks.update('H'+str(row+m)," Hrs Conv ")
            wks.update('I'+str(row+m),session["lesson"]+" ")
            wks.update('J'+str(row+m),session["remarks"]+" ")
            wks.update('K'+str(row+m)," ")
    row=row+m+2


#print(wks.acell('A9').value)
#print(wks.cell(3, 4).value)
#print(wks.get('A7:E9'))

#print(wks.get_all_records())
#print(wks.get_all_values())

wks.format("A1:K1", {"horizontalAlignment": "CENTER","textFormat": {"fontSize": 12,"bold": True}})

#wks.update('A1', 'Anthony')
#wks.update('D1:E2', [['Engineering', 'Tennis'], ['Business', 'Pottery']])
#wks.update('F1', '=UPPER(E2)', raw=False)

#wks.delete_rows(25)