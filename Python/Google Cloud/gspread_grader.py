import gspread
import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, colors

with open(os.path.join(os.path.dirname(__file__),"classesdone.json"), 'r') as f:
    data = json.load(f)
cred_name = os.path.join(os.path.dirname(__file__),'drive_serviceacct.json')

sa = gspread.service_account(filename=cred_name)
sh = sa.open("Grade 1 Section F (Responses)")

wks = sa.open("Grade 1 Section F (Responses)").sheet1

try:
    insertcol=["grade"]
    wks.insert_cols(insertcol,5)
except:
    pass
row=2
wks.update('E'+str(1),"Grade")
while True:
    
    score= wks.cell(row,4).value
    if score == None:
        break
    score_spl=score.split("/")
    score_val=int(score_spl[0])
    print(score_spl[0])

    if score_val >= 18:
        wks.update('E'+str(row),"A+")
    elif score_val > 15:
        wks.update('E'+str(row),"A")
    elif score_val > 12:
        wks.update('E'+str(row),"B+")
    elif score_val > 10:
        wks.update('E'+str(row),"B")
    elif score_val > 7:
        wks.update('E'+str(row),"C+")
    elif score_val > 5:
        wks.update('E'+str(row),"C")
    elif score_val > 2:
        wks.update('E'+str(row),"D+")
    else:
        wks.update('E'+str(row),"D")
        
    row=row+1