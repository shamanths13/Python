from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, colors
import os
import json

with open(os.path.join(os.path.dirname(__file__),"classesdone.json"), 'r') as f:
    data = json.load(f)

wb=Workbook()
ws=wb.active

row=1

for n,day in enumerate(data):
    ws['A'+str(row)].value=data[n]["day"]
    ws['A'+str(row)].font=Font(bold=True, size=12)
    ws['A'+str(row)].alignment=Alignment(vertical='center', horizontal='center')
    ws['B'+str(row)].value=data[n]["date"]
    ws['B'+str(row)].font=Font(bold=True, size=12 )
    ws['B'+str(row)].alignment=Alignment(vertical='center', horizontal='center')
    if data[n]["day"]=="Sunday":
        ws['A'+str(row)].font=Font(bold=True, size=12, color='FF0000')
        ws['B'+str(row)].font=Font(bold=True, size=12, color='FF0000')
    m=0
    if len(data[n]["classes"]) > 0:
        for m,session in enumerate(data[n]["classes"]):
            ws['C'+str(row+m)].value=session["section"]
            ws['D'+str(row+m)].value=session["cond"]
            ws['E'+str(row+m)].value=session["sch_time"]
            ws['F'+str(row+m)].value=session["act_time"]
            ws['G'+str(row+m)].value=session["hours"]
            if session["hours"] != "  Hours  ":
                if session["hours"]!= "":
                    hrs_red=session["hours"].split(":")
                    hrs_red=round((int(hrs_red[0])+(int(hrs_red[1])/60)),3)
                    ws['H'+str(row+m)].value=hrs_red
            else:
                ws['H'+str(row+m)].value=" Hrs Conv "
            ws['I'+str(row+m)].value=session["lesson"]+" "
            ws['J'+str(row+m)].value=session["remarks"]+" "
            ws['K'+str(row+m)].value=" "
    ws.merge_cells("A"+str(row)+":"+"A"+str(row+m))
    ws.merge_cells("B"+str(row)+":"+"B"+str(row+m))
    row=row+m+2

for col in range(1, 11):
    ws[get_column_letter(col)+'1'].font=Font(bold=True, size=12 )
    ws.column_dimensions[get_column_letter(col)].width = len(ws[get_column_letter(col)+'1'].value)*1.1
    ws[get_column_letter(col)+'1'].alignment=Alignment(horizontal='center')


def classes_sorted():

    global row

    classes_list=[]
    for n in range(len(data)-1):
        if len(data[n+1]["classes"]) > 0:
            for m,session in enumerate(data[n+1]["classes"]):
                class_name=data[n+1]["classes"][m]["section"]
                if class_name not in classes_list:
                    classes_list.append(class_name)

    classes_list=sorted(classes_list)
    ws['L'+'1'].value=" SUMMARY "
    ws['L'+'1'].font=Font(bold=True, size=12 )
    ws['L'+'1'].alignment=Alignment(horizontal='center')
    ws.merge_cells("L1"+":"+"O1")
    

    ws['L'+'2'].value="      Grade/Section      "
    ws['L'+'2'].font=Font(bold=True, size=12 )
    ws.column_dimensions['L'].width = len(ws['L'+'2'].value)*1.1
    ws['L'+'2'].alignment=Alignment(horizontal='center')

    ws['M'+'2'].value=" No. "
    ws['M'+'2'].font=Font(bold=True, size=12 )
    ws.column_dimensions['M'].width = len(ws['M'+'2'].value)*1.1
    ws['M'+'2'].alignment=Alignment(horizontal='center')

    ws['N'+'2'].value="  Hours  "
    ws['N'+'2'].font=Font(bold=True, size=12 )
    ws.column_dimensions['N'].width = len(ws['N'+'2'].value)*1.1
    ws['N'+'2'].alignment=Alignment(horizontal='center')

    ws['O'+'2'].value=" Hrs. Decimal "
    ws['O'+'2'].font=Font(bold=True, size=12 )
    ws.column_dimensions['O'].width = len(ws['O'+'2'].value)*1.1
    ws['O'+'2'].alignment=Alignment(horizontal='center')

    total_classes=0
    total_hrs_dec=0
    hrs_total=0
    min_total=0

    row=3
    for n, name_class in enumerate(classes_list):
        classes_no=0
        hrs_dec=0
        total_min=[]
        for n in range(len(data)):
            if len(data[n]["classes"]) > 0:
                for m,session in enumerate(data[n]["classes"]):
                    class_name=data[n]["classes"][m]["section"]
                    if (class_name == name_class) and (data[n]["classes"][m]["cond"] == "Y"):
                        classes_no=classes_no+1
                        hrs_red=session["hours"].split(":")
                        hrs_red=round((int(hrs_red[0])+(int(hrs_red[1])/60)),3)
                        hrs_dec=hrs_dec+hrs_red
                        pass
        total_min=str(hrs_dec).split(".")
        if len(total_min)>1:
            time_hrs=total_min[0]+":"+'{:02d}'.format(round(float("0."+total_min[1])*60))
        else:
            time_hrs=total_min[0]+":00"

        ws['L'+str(row)].value=name_class
        ws['M'+str(row)].value=classes_no
        ws['M'+str(row)].alignment=Alignment(horizontal='center')
        ws['N'+str(row)].value=time_hrs
        ws['N'+str(row)].alignment=Alignment(horizontal='right')
        ws['O'+str(row)].value=round(hrs_dec,3)

        total_classes=total_classes+classes_no
        total_hrs_dec=total_hrs_dec+hrs_dec
        hrs_min=time_hrs.split(":")
        hrs_total=hrs_total+int(hrs_min[0])
        min_total=min_total+int(hrs_min[1])

        row = row+1

    ws['L'+str(row)].value="Total: Rs. "+str(round((total_hrs_dec*500),2))
    ws['L'+str(row)].font=Font(bold=True, size=12 )
    ws['M'+str(row)].value=str(total_classes)
    ws['M'+str(row)].font=Font(bold=True, size=12 )
    ws['M'+str(row)].alignment=Alignment(horizontal='center')
    ws['N'+str(row)].value=str(int(hrs_total+((min_total-(min_total%60))/60)))+":"+'{:02d}'.format(round(int(min_total%60)))
    ws['N'+str(row)].font=Font(bold=True, size=12 )
    ws['N'+str(row)].alignment=Alignment(horizontal='right')
    ws['O'+str(row)].value=round(total_hrs_dec,3)
    ws['O'+str(row)].font=Font(bold=True, size=12 )

    row=row+3
    pass

def days_sorted():
    global row 

    ws['L'+str(row)].value=" Date "
    ws['L'+str(row)].font=Font(bold=True, size=12 )
    ws['L'+str(row)].alignment=Alignment(horizontal='center')

    ws['M'+str(row)].value=" No. "
    ws['M'+str(row)].font=Font(bold=True, size=12 )
    ws['M'+str(row)].alignment=Alignment(horizontal='center')

    row=row+1

    for n in range(len(data)-1):
        ws['L'+str(row)].value=data[n+1]["date"]+", "+data[n+1]["day"]
        sum_date_classes=0
        if len(data[n+1]["classes"])>0:
            for m in range(len(data[n+1]["classes"])):
                if data[n+1]["classes"][m]["cond"] == "Y":
                    sum_date_classes=sum_date_classes+1
                
        ws['M'+str(row)].value=sum_date_classes
        row=row+1

    pass

classes_sorted()
days_sorted()

wb.save(os.path.join(os.path.dirname(__file__),"CompletedClasses.xlsx"))

